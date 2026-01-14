import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from openpyxl.utils import get_column_letter

# Load the workbook
wb = openpyxl.load_workbook('excel/portpolio_app.xlsx')

# ============================================
# IMPROVE NETWORTH SHEET
# ============================================
ws_networth = wb['Networth']

# Insert new columns at the beginning and middle
# Column positions: A=ID, B=Platform, C=InvestmentType, D=TickerSymbol, E=AssetName, F=InvestedAmount, etc.

# Insert ID column at position A
ws_networth.insert_cols(1)
ws_networth.cell(1, 1, 'ID')

# The original columns are now shifted:
# A=ID, B=Platform, C=Investment Type, D=Invested Amount, E=Current Amount, F=Profit/Loss, G=Currency, H=Updated Date

# Insert new columns after Investment Type (position 4)
ws_networth.insert_cols(4)  # For Ticker Symbol
ws_networth.cell(1, 4, 'Ticker Symbol')

ws_networth.insert_cols(5)  # For Asset Name
ws_networth.cell(1, 5, 'Asset Name')

# Now columns are:
# A=ID, B=Platform, C=Investment Type, D=Ticker Symbol, E=Asset Name, F=Invested Amount, G=Current Amount, H=Profit/Loss, I=Currency, J=Updated Date

# Insert columns after Updated Date (now at position 10)
ws_networth.insert_cols(11)  # For Purchase Date
ws_networth.cell(1, 11, 'Purchase Date')

ws_networth.insert_cols(12)  # For Quantity
ws_networth.cell(1, 12, 'Quantity')

ws_networth.insert_cols(13)  # For Auto Update
ws_networth.cell(1, 13, 'Auto Update')

ws_networth.insert_cols(14)  # For Notes
ws_networth.cell(1, 14, 'Notes')

# Final columns:
# A=ID, B=Platform, C=Investment Type, D=Ticker Symbol, E=Asset Name, F=Invested Amount,
# G=Current Amount, H=Profit/Loss, I=Currency, J=Updated Date, K=Purchase Date,
# L=Quantity, M=Auto Update, N=Notes

# Generate IDs and populate new fields for existing data
row_num = 2
id_counter = 1

while True:
    platform_value = ws_networth.cell(row_num, 2).value  # Platform column

    # Stop if we hit empty rows
    if platform_value is None:
        break

    # Generate ID
    investment_id = f"INV{str(id_counter).zfill(3)}"
    ws_networth.cell(row_num, 1, investment_id)

    # Get investment type
    investment_type = ws_networth.cell(row_num, 3).value

    # Populate Asset Name from Platform (can be updated manually later)
    platform = ws_networth.cell(row_num, 2).value
    ws_networth.cell(row_num, 5, f"{platform} - {investment_type}")

    # Set Auto Update to YES for STOCK and CRYPTO, NO for others
    if investment_type in ['STOCK', 'CRYPTO']:
        ws_networth.cell(row_num, 13, 'YES')
        # Add placeholder for Ticker Symbol (to be filled manually)
        ws_networth.cell(row_num, 4, 'TO_BE_ADDED')
    else:
        ws_networth.cell(row_num, 13, 'NO')

    # Fix date format in Updated Date column (J)
    updated_date = ws_networth.cell(row_num, 10).value
    if isinstance(updated_date, str):
        try:
            # Try to parse string date formats like "27-08-2025"
            parsed_date = datetime.strptime(updated_date, "%d-%m-%Y")
            ws_networth.cell(row_num, 10, parsed_date)
        except:
            pass

    row_num += 1
    id_counter += 1

# Style the header row
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

for col in range(1, 15):
    cell = ws_networth.cell(1, col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Auto-adjust column widths
for col in range(1, 15):
    column_letter = get_column_letter(col)
    ws_networth.column_dimensions[column_letter].width = 15

# ============================================
# IMPROVE INVESTMENT RETURNS SHEET
# ============================================
ws_returns = wb['Investment Returns']

# Insert Investment ID column at position A
ws_returns.insert_cols(1)
ws_returns.cell(1, 1, 'Investment ID')

# Now columns are:
# A=Investment ID, B=Stock/Instrument, C=Investment Type/Return type, D=Date, E=Amount, F=Currency, G=Notes

# Rename column C to just "Return Type"
ws_returns.cell(1, 3, 'Return Type')

# Update column G header to "Notes"
ws_returns.cell(1, 7, 'Notes')

# Delete extra empty columns (H onwards - now columns 8+)
max_col = ws_returns.max_column
if max_col > 7:
    ws_returns.delete_cols(8, max_col - 7)

# Fix date formats and add placeholder Investment IDs
row_num = 2
while True:
    instrument = ws_returns.cell(row_num, 2).value  # Stock/Instrument column

    # Stop if we hit empty rows
    if instrument is None:
        break

    # Add placeholder for Investment ID (to be filled manually or via lookup)
    ws_returns.cell(row_num, 1, 'TO_BE_LINKED')

    # Fix date format in Date column (D)
    date_value = ws_returns.cell(row_num, 4).value
    if isinstance(date_value, float):
        # Year-only values like 2025.0 - set to first day of year
        try:
            year = int(date_value)
            ws_returns.cell(row_num, 4, datetime(year, 1, 1))
        except:
            pass
    elif isinstance(date_value, str):
        try:
            parsed_date = datetime.strptime(date_value, "%d-%m-%Y")
            ws_returns.cell(row_num, 4, parsed_date)
        except:
            pass

    row_num += 1

# Style the header row for Investment Returns
for col in range(1, 8):
    cell = ws_returns.cell(1, col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Auto-adjust column widths
for col in range(1, 8):
    column_letter = get_column_letter(col)
    ws_returns.column_dimensions[column_letter].width = 18

# ============================================
# CREATE NEW SHEET: STOCK SYMBOLS (OPTIONAL)
# ============================================
if 'Stock Symbols' not in wb.sheetnames:
    ws_stocks = wb.create_sheet('Stock Symbols')

    # Create headers
    headers = ['Ticker Symbol', 'Asset Name', 'Exchange', 'Currency', 'Auto Fetch', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws_stocks.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add sample data as examples
    sample_data = [
        ['AAPL', 'Apple Inc', 'NASDAQ', 'USD', 'YES', 'Example stock'],
        ['DBS.SI', 'DBS Bank', 'SGX', 'SGD', 'YES', 'Singapore stock'],
        ['BTC-USD', 'Bitcoin', 'CRYPTO', 'USD', 'YES', 'Cryptocurrency'],
    ]

    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws_stocks.cell(row_idx, col_idx, value)

    # Auto-adjust column widths
    for col in range(1, 7):
        column_letter = get_column_letter(col)
        ws_stocks.column_dimensions[column_letter].width = 18

# ============================================
# SAVE THE IMPROVED WORKBOOK
# ============================================
output_file = 'excel/portpolio_app_improved.xlsx'
wb.save(output_file)

print(f"✅ Excel file improved successfully!")
print(f"📁 Saved as: {output_file}")
print("\n📋 Changes made:")
print("\n🔹 Networth Sheet:")
print("  - Added 'ID' column with auto-generated IDs (INV001, INV002, ...)")
print("  - Added 'Ticker Symbol' column (marked as TO_BE_ADDED for stocks)")
print("  - Added 'Asset Name' column (auto-populated from Platform + Type)")
print("  - Added 'Purchase Date' column (empty - to be filled)")
print("  - Added 'Quantity' column (empty - to be filled for stocks)")
print("  - Added 'Auto Update' column (YES for stocks/crypto, NO for others)")
print("  - Added 'Notes' column for additional information")
print("  - Fixed date formats where possible")
print("  - Styled headers with blue background")
print("\n🔹 Investment Returns Sheet:")
print("  - Added 'Investment ID' column (marked as TO_BE_LINKED)")
print("  - Renamed 'Investment Type/Return type' to 'Return Type'")
print("  - Removed all empty columns")
print("  - Added 'Notes' column")
print("  - Fixed date formats (year-only values set to Jan 1)")
print("  - Styled headers")
print("\n🔹 Stock Symbols Sheet (NEW):")
print("  - Created new sheet for managing stock ticker symbols")
print("  - Added sample entries as examples")
print("\n📝 Next Steps:")
print("  1. Open the improved file: excel/portpolio_app_improved.xlsx")
print("  2. Fill in 'Ticker Symbol' for all STOCK entries (replace TO_BE_ADDED)")
print("  3. Update 'Asset Name' to more descriptive names if needed")
print("  4. Add 'Purchase Date' for your investments")
print("  5. Add 'Quantity' for stock holdings")
print("  6. Link 'Investment ID' in Investment Returns sheet to match Networth IDs")
print("  7. Review and verify all data")
