import openpyxl

wb = openpyxl.load_workbook('excel/portpolio_app_improved.xlsx')

print("=" * 80)
print("NETWORTH SHEET - First 5 Rows")
print("=" * 80)
ws1 = wb['Networth']

# Print headers
headers = []
for col in range(1, ws1.max_column + 1):
    headers.append(ws1.cell(1, col).value or '')

print(f"\n{' | '.join(headers)}")
print("-" * 80)

# Print first 5 data rows
for row in range(2, min(7, ws1.max_row + 1)):
    row_data = []
    for col in range(1, ws1.max_column + 1):
        cell_value = ws1.cell(row, col).value
        if cell_value is None:
            row_data.append('')
        else:
            row_data.append(str(cell_value)[:20])  # Limit length
    print(f"{' | '.join(row_data)}")

print("\n" + "=" * 80)
print("INVESTMENT RETURNS SHEET - First 5 Rows")
print("=" * 80)
ws2 = wb['Investment Returns']

# Print headers
headers = []
for col in range(1, ws2.max_column + 1):
    headers.append(ws2.cell(1, col).value or '')

print(f"\n{' | '.join(headers)}")
print("-" * 80)

# Print first 5 data rows
for row in range(2, min(7, ws2.max_row + 1)):
    row_data = []
    for col in range(1, ws2.max_column + 1):
        cell_value = ws2.cell(row, col).value
        if cell_value is None:
            row_data.append('')
        else:
            row_data.append(str(cell_value)[:25])  # Limit length
    print(f"{' | '.join(row_data)}")

print("\n" + "=" * 80)
print("STOCK SYMBOLS SHEET - Sample Data")
print("=" * 80)
ws3 = wb['Stock Symbols']

# Print all data
for row in ws3.iter_rows(values_only=True):
    print(' | '.join(str(cell) if cell else '' for cell in row))

print("\n" + "=" * 80)
print("SUMMARY")
print("=" * 80)
print(f"Total sheets: {len(wb.sheetnames)}")
print(f"Sheet names: {', '.join(wb.sheetnames)}")
print(f"\nNetworth entries: {ws1.max_row - 1} rows")
print(f"Investment Returns entries: {ws2.max_row - 1} rows")
print(f"Stock Symbols entries: {ws3.max_row - 1} rows")
