import openpyxl

wb = openpyxl.load_workbook('excel/portpolio_app.xlsx')
print('Sheet names:', wb.sheetnames)

print('\n=== Networth Sheet ===')
ws1 = wb['Networth']
for row in ws1.iter_rows(values_only=True):
    print(row)

print('\n=== Investment Returns Sheet ===')
ws2 = wb['Investment Returns']
for row in ws2.iter_rows(values_only=True):
    print(row)
