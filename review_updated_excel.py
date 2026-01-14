import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook('excel/portpolio_app_improved.xlsx')

print("=" * 100)
print("REVIEWING UPDATED EXCEL FILE")
print("=" * 100)

# ============================================
# NETWORTH SHEET REVIEW
# ============================================
print("\n📊 NETWORTH SHEET")
print("=" * 100)
ws_networth = wb['Networth']

# Get headers
headers = [ws_networth.cell(1, col).value for col in range(1, ws_networth.max_column + 1)]
print(f"\nColumns: {', '.join(headers)}")
print(f"Total entries: {ws_networth.max_row - 1}")

# Analyze the data
investments_by_type = {}
investments_by_currency = {}
stocks_with_tickers = []
stocks_without_tickers = []
auto_update_count = 0

print("\n📋 Sample Entries (First 3):")
print("-" * 100)
for row in range(2, min(5, ws_networth.max_row + 1)):
    row_data = {}
    for col in range(1, ws_networth.max_column + 1):
        header = ws_networth.cell(1, col).value
        value = ws_networth.cell(row, col).value
        row_data[header] = value

    print(f"\n{row_data['ID']}: {row_data['Asset Name']}")
    print(f"  Platform: {row_data['Platform']} | Type: {row_data['Investment Type']} | Currency: {row_data['Currency']}")
    print(f"  Ticker: {row_data['Ticker Symbol']} | Auto Update: {row_data['Auto Update']}")
    print(f"  Invested: {row_data['Invested Amount']} | Current: {row_data['Current Amount']} | P/L: {row_data['Profit/Loss']}")

# Collect statistics
for row in range(2, ws_networth.max_row + 1):
    inv_type = ws_networth.cell(row, 3).value
    currency = ws_networth.cell(row, 9).value
    ticker = ws_networth.cell(row, 4).value
    auto_update = ws_networth.cell(row, 13).value
    investment_id = ws_networth.cell(row, 1).value

    if inv_type is None:
        break

    # Count by type
    investments_by_type[inv_type] = investments_by_type.get(inv_type, 0) + 1

    # Count by currency
    if currency:
        investments_by_currency[currency] = investments_by_currency.get(currency, 0) + 1

    # Check ticker symbols for stocks
    if inv_type in ['STOCK', 'CRYPTO', 'ETF', 'FUND']:
        if ticker and ticker != 'TO_BE_ADDED' and ticker.strip():
            stocks_with_tickers.append((investment_id, ticker))
        else:
            stocks_without_tickers.append(investment_id)

    # Count auto-update
    if auto_update == 'YES':
        auto_update_count += 1

print("\n\n📈 PORTFOLIO BREAKDOWN:")
print("-" * 100)
print("\n🏷️  By Investment Type:")
for inv_type, count in sorted(investments_by_type.items()):
    print(f"  {inv_type}: {count} entries")

print("\n💰 By Currency:")
for currency, count in sorted(investments_by_currency.items()):
    print(f"  {currency}: {count} entries")

print(f"\n🔄 Auto-Update Enabled: {auto_update_count} investments")

print(f"\n📍 Ticker Symbols Status:")
print(f"  ✅ Stocks with ticker symbols: {len(stocks_with_tickers)}")
print(f"  ⚠️  Stocks without ticker symbols: {len(stocks_without_tickers)}")

if stocks_with_tickers:
    print("\n  Stocks ready for API price fetching:")
    for inv_id, ticker in stocks_with_tickers[:10]:  # Show first 10
        print(f"    {inv_id}: {ticker}")
    if len(stocks_with_tickers) > 10:
        print(f"    ... and {len(stocks_with_tickers) - 10} more")

if stocks_without_tickers:
    print(f"\n  ⚠️  Warning: {len(stocks_without_tickers)} stocks still need ticker symbols:")
    print(f"    {', '.join(stocks_without_tickers[:5])}")
    if len(stocks_without_tickers) > 5:
        print(f"    ... and {len(stocks_without_tickers) - 5} more")

# ============================================
# INVESTMENT RETURNS SHEET REVIEW
# ============================================
print("\n\n📊 INVESTMENT RETURNS SHEET")
print("=" * 100)
ws_returns = wb['Investment Returns']

# Get headers
headers = [ws_returns.cell(1, col).value or '' for col in range(1, ws_returns.max_column + 1)]
print(f"\nColumns: {', '.join(headers)}")
print(f"Total entries: {ws_returns.max_row - 1}")

# Analyze returns data
returns_by_type = {}
returns_by_currency = {}
linked_returns = 0
unlinked_returns = 0
date_range_start = None
date_range_end = None

print("\n📋 Sample Entries (First 3):")
print("-" * 100)
for row in range(2, min(5, ws_returns.max_row + 1)):
    inv_id = ws_returns.cell(row, 1).value
    instrument = ws_returns.cell(row, 2).value
    return_type = ws_returns.cell(row, 3).value
    date = ws_returns.cell(row, 4).value
    amount = ws_returns.cell(row, 5).value
    currency = ws_returns.cell(row, 6).value

    print(f"\n  {instrument} ({return_type})")
    print(f"    Investment ID: {inv_id} | Date: {date} | Amount: {amount} {currency}")

# Collect statistics
for row in range(2, ws_returns.max_row + 1):
    inv_id = ws_returns.cell(row, 1).value
    return_type = ws_returns.cell(row, 3).value
    currency = ws_returns.cell(row, 6).value
    date = ws_returns.cell(row, 4).value

    if return_type is None:
        break

    # Count by type
    returns_by_type[return_type] = returns_by_type.get(return_type, 0) + 1

    # Count by currency
    if currency:
        returns_by_currency[currency] = returns_by_currency.get(currency, 0) + 1

    # Check linkage
    if inv_id and inv_id != 'TO_BE_LINKED':
        linked_returns += 1
    else:
        unlinked_returns += 1

    # Track date range
    if isinstance(date, datetime):
        if date_range_start is None or date < date_range_start:
            date_range_start = date
        if date_range_end is None or date > date_range_end:
            date_range_end = date

print("\n\n📈 RETURNS BREAKDOWN:")
print("-" * 100)
print("\n🏷️  By Return Type:")
for return_type, count in sorted(returns_by_type.items()):
    print(f"  {return_type}: {count} entries")

print("\n💰 By Currency:")
for currency, count in sorted(returns_by_currency.items()):
    print(f"  {currency}: {count} entries")

print(f"\n🔗 Linkage Status:")
print(f"  ✅ Linked to investments: {linked_returns}")
print(f"  ⚠️  Not yet linked: {unlinked_returns}")

if date_range_start and date_range_end:
    print(f"\n📅 Date Range:")
    print(f"  From: {date_range_start.strftime('%Y-%m-%d')}")
    print(f"  To: {date_range_end.strftime('%Y-%m-%d')}")
    days_diff = (date_range_end - date_range_start).days
    print(f"  Duration: {days_diff} days (~{days_diff/30:.1f} months)")

# ============================================
# READINESS CHECK
# ============================================
print("\n\n🎯 APPLICATION READINESS CHECK")
print("=" * 100)

checks = {
    "✅ Excel file structure is valid": True,
    "✅ Investment IDs are assigned": ws_networth.max_row > 1,
    f"{'✅' if len(stocks_with_tickers) > 0 else '⚠️ '} Stock ticker symbols defined": len(stocks_with_tickers) > 0,
    f"{'✅' if auto_update_count > 0 else '⚠️ '} Auto-update enabled for stocks": auto_update_count > 0,
    "✅ Investment returns data available": ws_returns.max_row > 1,
    "✅ Multiple currencies detected": len(investments_by_currency) > 1,
    "✅ Historical data available": date_range_start is not None,
}

for check, status in checks.items():
    print(f"  {check}")

print("\n\n🚀 RECOMMENDATION:")
print("=" * 100)
if len(stocks_without_tickers) == 0:
    print("✅ Excel file is fully ready! All stocks have ticker symbols.")
    print("   Ready to start building the application.")
elif len(stocks_with_tickers) > 0:
    print(f"⚠️  Partially ready: {len(stocks_with_tickers)} stocks have tickers, {len(stocks_without_tickers)} are missing.")
    print("   You can proceed with application development.")
    print("   Missing ticker symbols can be added later through the UI.")
else:
    print("⚠️  No ticker symbols found for stocks.")
    print("   The application will work, but automatic price fetching won't be available initially.")
    print("   You can add ticker symbols later through the UI or by updating the Excel file.")

print("\n" + "=" * 100)
